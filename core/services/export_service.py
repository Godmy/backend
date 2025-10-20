"""
Export Service - экспорт данных в различные форматы
"""

import json
import csv
import os
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from io import StringIO, BytesIO

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

from core.models.import_export_job import (
    ImportExportJobModel,
    JobStatus,
    JobType,
    ExportFormat,
    EntityType,
)
from languages.models.concept import ConceptModel
from languages.models.dictionary import DictionaryModel
from languages.models.language import LanguageModel
from auth.models.user import UserModel


class ExportService:
    """Сервис для экспорта данных"""

    def __init__(self, db: Session):
        self.db = db
        self.export_dir = os.getenv("EXPORT_DIR", "exports")
        os.makedirs(self.export_dir, exist_ok=True)

    def create_export_job(
        self,
        user_id: int,
        entity_type: EntityType,
        format: ExportFormat,
        filters: Optional[Dict[str, Any]] = None,
    ) -> ImportExportJobModel:
        """Создать задание экспорта"""
        job = ImportExportJobModel(
            user_id=user_id,
            job_type=JobType.EXPORT,
            entity_type=entity_type,
            format=format,
            status=JobStatus.PENDING,
            filters=filters or {},
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def process_export(self, job_id: int) -> ImportExportJobModel:
        """Обработать экспорт (синхронно)"""
        job = self.db.query(ImportExportJobModel).filter_by(id=job_id).first()
        if not job:
            raise ValueError(f"Job {job_id} not found")

        try:
            job.status = JobStatus.PROCESSING
            self.db.commit()

            # Получаем данные в зависимости от типа сущности
            if job.entity_type == EntityType.CONCEPTS:
                data = self._export_concepts(job.filters)
            elif job.entity_type == EntityType.DICTIONARIES:
                data = self._export_dictionaries(job.filters)
            elif job.entity_type == EntityType.USERS:
                data = self._export_users(job.filters)
            elif job.entity_type == EntityType.LANGUAGES:
                data = self._export_languages(job.filters)
            else:
                raise ValueError(f"Unknown entity type: {job.entity_type}")

            job.total_count = len(data)

            # Экспортируем в нужный формат
            if job.format == ExportFormat.JSON:
                file_path = self._export_to_json(data, job)
            elif job.format == ExportFormat.CSV:
                file_path = self._export_to_csv(data, job)
            elif job.format == ExportFormat.XLSX:
                file_path = self._export_to_xlsx(data, job)
            else:
                raise ValueError(f"Unknown format: {job.format}")

            # Обновляем job
            job.file_url = f"/exports/{os.path.basename(file_path)}"
            job.expires_at = int((datetime.utcnow() + timedelta(hours=24)).timestamp())
            job.processed_count = len(data)
            job.status = JobStatus.COMPLETED
            self.db.commit()
            self.db.refresh(job)

            return job

        except Exception as e:
            job.status = JobStatus.FAILED
            job.error_message = str(e)
            self.db.commit()
            raise

    def _export_concepts(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Экспорт концепций"""
        query = self.db.query(ConceptModel).filter(ConceptModel.deleted_at.is_(None))

        # Применяем фильтры
        if filters.get("language"):
            # Фильтр по языку через dictionaries
            lang_code = filters["language"]
            query = query.join(DictionaryModel).join(LanguageModel).filter(
                LanguageModel.code == lang_code
            )

        concepts = query.all()

        result = []
        for concept in concepts:
            concept_data = {
                "id": concept.id,
                "parent_id": concept.parent_id,
                "path": concept.path,
                "depth": concept.depth,
                "created_at": concept.created_at.isoformat() if concept.created_at else None,
                "updated_at": concept.updated_at.isoformat() if concept.updated_at else None,
                "translations": [],
            }

            # Добавляем переводы
            for dictionary in concept.dictionaries:
                if dictionary.deleted_at is None:
                    concept_data["translations"].append(
                        {
                            "language_code": dictionary.language.code,
                            "language_name": dictionary.language.name,
                            "name": dictionary.name,
                            "description": dictionary.description,
                            "image": dictionary.image,
                        }
                    )

            result.append(concept_data)

        return result

    def _export_dictionaries(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Экспорт словарей"""
        query = self.db.query(DictionaryModel).filter(DictionaryModel.deleted_at.is_(None))

        # Применяем фильтры
        if filters.get("language"):
            lang_code = filters["language"]
            query = query.join(LanguageModel).filter(LanguageModel.code == lang_code)

        dictionaries = query.all()

        result = []
        for dictionary in dictionaries:
            result.append(
                {
                    "id": dictionary.id,
                    "concept_id": dictionary.concept_id,
                    "concept_path": dictionary.concept.path if dictionary.concept else None,
                    "language_id": dictionary.language_id,
                    "language_code": dictionary.language.code,
                    "language_name": dictionary.language.name,
                    "name": dictionary.name,
                    "description": dictionary.description,
                    "image": dictionary.image,
                    "created_at": dictionary.created_at.isoformat()
                    if dictionary.created_at
                    else None,
                    "updated_at": dictionary.updated_at.isoformat()
                    if dictionary.updated_at
                    else None,
                }
            )

        return result

    def _export_users(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Экспорт пользователей (без паролей)"""
        query = self.db.query(UserModel).filter(UserModel.deleted_at.is_(None))

        # Применяем фильтры
        if filters.get("role"):
            # TODO: фильтр по ролям (требует join с UserRoleModel)
            pass

        users = query.all()

        result = []
        for user in users:
            user_data = {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "is_active": user.is_active,
                "is_verified": user.is_verified,
                "created_at": user.created_at.isoformat() if user.created_at else None,
                "updated_at": user.updated_at.isoformat() if user.updated_at else None,
                "roles": [],
            }

            # Добавляем роли
            for user_role in user.roles:
                user_data["roles"].append(user_role.role.name)

            # Добавляем профиль
            if user.profile:
                user_data["profile"] = {
                    "first_name": user.profile.first_name,
                    "last_name": user.profile.last_name,
                    "bio": user.profile.bio,
                }

            result.append(user_data)

        return result

    def _export_languages(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Экспорт языков"""
        query = self.db.query(LanguageModel).filter(LanguageModel.deleted_at.is_(None))

        languages = query.all()

        result = []
        for lang in languages:
            result.append(
                {
                    "id": lang.id,
                    "code": lang.code,
                    "name": lang.name,
                    "native_name": lang.native_name,
                    "rtl": lang.rtl,
                    "created_at": lang.created_at.isoformat() if lang.created_at else None,
                    "updated_at": lang.updated_at.isoformat() if lang.updated_at else None,
                }
            )

        return result

    def _export_to_json(self, data: List[Dict[str, Any]], job: ImportExportJobModel) -> str:
        """Экспорт в JSON"""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{job.entity_type.value}_{timestamp}.json"
        file_path = os.path.join(self.export_dir, filename)

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return file_path

    def _export_to_csv(self, data: List[Dict[str, Any]], job: ImportExportJobModel) -> str:
        """Экспорт в CSV (плоская структура)"""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{job.entity_type.value}_{timestamp}.csv"
        file_path = os.path.join(self.export_dir, filename)

        if not data:
            # Создаем пустой файл
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("")
            return file_path

        # Flatten nested structures для CSV
        flat_data = self._flatten_data(data)

        # Используем pandas для удобства
        df = pd.DataFrame(flat_data)
        df.to_csv(file_path, index=False, encoding="utf-8")

        return file_path

    def _export_to_xlsx(self, data: List[Dict[str, Any]], job: ImportExportJobModel) -> str:
        """Экспорт в XLSX"""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{job.entity_type.value}_{timestamp}.xlsx"
        file_path = os.path.join(self.export_dir, filename)

        # Flatten data
        flat_data = self._flatten_data(data)

        if not flat_data:
            # Создаем пустой workbook
            wb = Workbook()
            ws = wb.active
            ws.title = job.entity_type.value
            wb.save(file_path)
            return file_path

        # Используем pandas для создания XLSX
        df = pd.DataFrame(flat_data)

        # Создаем writer
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=job.entity_type.value, index=False)

            # Форматирование
            worksheet = writer.sheets[job.entity_type.value]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

        return file_path

    def _flatten_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Flatten вложенных структур для CSV/XLSX"""
        result = []

        for item in data:
            flat_item = {}

            for key, value in item.items():
                if isinstance(value, list):
                    # Преобразуем списки в строки
                    if value and isinstance(value[0], dict):
                        flat_item[key] = json.dumps(value, ensure_ascii=False)
                    else:
                        flat_item[key] = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    # Flatten словари
                    for nested_key, nested_value in value.items():
                        flat_item[f"{key}_{nested_key}"] = nested_value
                else:
                    flat_item[key] = value

            result.append(flat_item)

        return result

    def get_export_file_path(self, job_id: int) -> Optional[str]:
        """Получить путь к экспортированному файлу"""
        job = self.db.query(ImportExportJobModel).filter_by(id=job_id).first()
        if not job or not job.file_url:
            return None

        filename = os.path.basename(job.file_url)
        file_path = os.path.join(self.export_dir, filename)

        if os.path.exists(file_path):
            return file_path

        return None

    def cleanup_old_exports(self, hours: int = 24) -> int:
        """Удалить старые экспортированные файлы"""
        cutoff_time = datetime.utcnow() - timedelta(hours=hours)
        deleted_count = 0

        for filename in os.listdir(self.export_dir):
            file_path = os.path.join(self.export_dir, filename)
            file_time = datetime.fromtimestamp(os.path.getmtime(file_path))

            if file_time < cutoff_time:
                try:
                    os.remove(file_path)
                    deleted_count += 1
                except Exception:
                    pass

        return deleted_count
